"""Strict Excel export: one file per video. A1=URL, from row2: A=domain, B=comment."""

from openpyxl import Workbook

def write_video_excel(filepath: str, video_url: str, domain: str, comments: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = video_url

    row = 2
    for c in comments:
        ws.cell(row=row, column=1).value = domain
        ws.cell(row=row, column=2).value = c
        row += 1

    wb.save(filepath)
