"""Report current actual comment counts per domain from deliverables or filtered data."""

from __future__ import annotations
import json
from datetime import datetime, timezone
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

from config import DELIVERABLES_DIR, YOUTUBE_COMMENTS_FILTERED_DIR, ANALYTICS_DIR, DOMAINS

def count_excel_comments(path: Path) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        val = row[0]
        if val is not None and str(val).strip():
            count += 1
    wb.close()
    return count

def count_jsonl_comments(path: Path) -> int:
    count = 0
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                text = obj.get("comment") or obj.get("text") or ""
                if str(text).strip():
                    count += 1
            except json.JSONDecodeError:
                if line:
                    count += 1
    return count

def count_csv_comments(path: Path) -> int:
    df = pd.read_csv(path)
    for col in ["comment", "text"]:
        if col in df.columns:
            return int(df[col].dropna().astype(str).str.strip().ne("").sum())
    return int(len(df))

def scan_deliverables():
    per_domain = {d: {"videos_with_data": 0, "total_saved_comments": 0, "videos_total": 0} for d in DOMAINS}
    per_video = []
    latest_mtime = None
    found = False

    for domain in DOMAINS:
        domain_dir = DELIVERABLES_DIR / domain
        if not domain_dir.exists():
            continue
        for xlsx in domain_dir.glob("*.xlsx"):
            found = True
            count = count_excel_comments(xlsx)
            per_domain[domain]["videos_total"] += 1
            if count > 0:
                per_domain[domain]["videos_with_data"] += 1
            per_domain[domain]["total_saved_comments"] += count
            per_video.append({
                "domain": domain,
                "video_id": xlsx.stem,
                "saved_to_excel_count": count,
                "excel_path": str(xlsx),
            })
            mtime = xlsx.stat().st_mtime
            latest_mtime = mtime if latest_mtime is None else max(latest_mtime, mtime)

    return found, per_domain, per_video, latest_mtime

def scan_filtered_comments():
    per_domain = {d: {"videos_with_data": 0, "total_saved_comments": 0, "videos_total": 0} for d in DOMAINS}
    per_video = []
    latest_mtime = None
    found = False

    for domain in DOMAINS:
        domain_dir = YOUTUBE_COMMENTS_FILTERED_DIR / domain
        if not domain_dir.exists():
            continue
        for path in domain_dir.glob("*"):
            if path.suffix not in [".jsonl", ".csv"]:
                continue
            found = True
            if path.suffix == ".jsonl":
                count = count_jsonl_comments(path)
            else:
                count = count_csv_comments(path)
            per_domain[domain]["videos_total"] += 1
            if count > 0:
                per_domain[domain]["videos_with_data"] += 1
            per_domain[domain]["total_saved_comments"] += count
            per_video.append({
                "domain": domain,
                "video_id": path.stem,
                "saved_to_excel_count": count,
                "file_path": str(path),
            })
            mtime = path.stat().st_mtime
            latest_mtime = mtime if latest_mtime is None else max(latest_mtime, mtime)

    return found, per_domain, per_video, latest_mtime

def main():
    ANALYTICS_DIR.mkdir(parents=True, exist_ok=True)
    by_video_path = ANALYTICS_DIR / "comments_progress_by_video.csv"
    by_domain_path = ANALYTICS_DIR / "comments_progress_by_domain.json"

    if not by_video_path.exists():
        df = pd.DataFrame(columns=[
            "timestamp",
            "domain",
            "video_id",
            "fetched_raw_count",
            "az_accepted_count",
            "saved_to_excel_count",
            "excel_path",
            "status",
            "error_message",
        ])
        df.to_csv(by_video_path, index=False, encoding="utf-8")

    if not by_domain_path.exists():
        payload = {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "domains": {
                d: {
                    "videos_processed": 0,
                    "videos_with_data": 0,
                    "raw_fetched_total": 0,
                    "az_accepted_total": 0,
                    "saved_to_excel_total": 0,
                    "target_10000_reached": False,
                } for d in DOMAINS
            }
        }
        by_domain_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    found_deliverables, per_domain, per_video, latest_mtime = scan_deliverables()
    source = "deliverables"

    if not found_deliverables:
        found_filtered, per_domain, per_video, latest_mtime = scan_filtered_comments()
        source = "filtered_comments" if found_filtered else "none"

    summary_rows = []
    for domain in DOMAINS:
        summary_rows.append({
            "domain": domain,
            "videos_with_data": per_domain[domain]["videos_with_data"],
            "total_saved_comments": per_domain[domain]["total_saved_comments"],
        })

    df = pd.DataFrame(summary_rows)
    print(df.to_string(index=False))

    if source == "none":
        print("No deliverables found yet; only estimates exist.")

    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": source,
        "last_modified": datetime.fromtimestamp(latest_mtime, timezone.utc).isoformat() if latest_mtime else None,
        "per_domain": per_domain,
        "per_video": per_video,
    }

    (ANALYTICS_DIR / "comments_progress_current.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    df.to_csv(ANALYTICS_DIR / "comments_progress_current.csv", index=False, encoding="utf-8")

if __name__ == "__main__":
    main()
